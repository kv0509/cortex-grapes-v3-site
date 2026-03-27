#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import math
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "Data"
REPORTS_DIR = ROOT / "Reports"
OUTPUT_PATH = ROOT / "dashboard" / "data" / "joint_backtest_overview.json"

PORTFOLIO_PATH = REPORTS_DIR / "portfolio_sequential.csv"
BTC_CANDLES_PATH = DATA_DIR / "BTCv1.csv"
ETH_CANDLES_PATH = DATA_DIR / "ETHv1.csv"
SOL_CANDLES_PATH = DATA_DIR / "SOLv1.csv"
REGIME_AUDIT_PATH = REPORTS_DIR / "regime_trade_audit.xlsx"

INITIAL_EQUITY = 6_000.0
REGIME_BUFFER_PCT = 0.002
TARGET_MARGIN = 20.0
TARGET_LEVERAGE = 100.0
TARGET_NOMINAL = TARGET_MARGIN * TARGET_LEVERAGE


def parse_dt(value: Any) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace(" ", "T"))
    except ValueError:
        return None


def to_ts(dt_value: datetime) -> str:
    return dt_value.strftime("%Y-%m-%d %H:%M:%S")


def parse_float(value: Any, default: float = 0.0) -> float:
    text = str(value or "").strip()
    if not text:
        return default
    return float(text)


def round_value(value: float, digits: int = 4) -> float:
    return round(float(value), digits)


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def load_portfolio_trades() -> List[Dict[str, Any]]:
    trades: List[Dict[str, Any]] = []
    for row in read_csv_rows(PORTFOLIO_PATH):
        exit_dt = parse_dt(row.get("Exit_Time"))
        entry_dt = parse_dt(row.get("Entry_Time"))
        if exit_dt is None or entry_dt is None:
            continue

        source_margin = parse_float(row.get("Margin"))
        source_nominal = parse_float(row.get("Nominal"))
        scale = TARGET_NOMINAL / source_nominal if source_nominal > 0 else 1.0

        trades.append(
            {
                "asset": str(row.get("Asset") or "").upper(),
                "direction": str(row.get("Direction") or "").title(),
                "entry_ts": to_ts(entry_dt),
                "exit_ts": to_ts(exit_dt),
                "entry_dt": entry_dt,
                "exit_dt": exit_dt,
                "entry_price": parse_float(row.get("Entry_Price")),
                "exit_price": parse_float(row.get("Exit_Price")),
                "margin": TARGET_MARGIN,
                "nominal": TARGET_NOMINAL,
                "gross_pnl": parse_float(row.get("Gross_PnL")) * scale,
                "fees": parse_float(row.get("Fees")) * scale,
                "net_pnl": parse_float(row.get("Net_PnL")) * scale,
                "type": str(row.get("Type") or "Unknown"),
                "source_margin": source_margin,
                "source_nominal": source_nominal,
                "scale_factor": scale,
            }
        )
    trades.sort(key=lambda item: (item["exit_dt"], item["asset"]))
    return trades


def load_btc_candles() -> List[Dict[str, Any]]:
    candles: List[Dict[str, Any]] = []
    ema144: Optional[float] = None
    ema169: Optional[float] = None
    alpha144 = 2.0 / (144.0 + 1.0)
    alpha169 = 2.0 / (169.0 + 1.0)

    for row in read_csv_rows(BTC_CANDLES_PATH):
        dt_value = parse_dt(row.get("dt"))
        close_value = parse_float(row.get("close"), default=float("nan"))
        if dt_value is None or math.isnan(close_value):
            continue

        if ema144 is None:
            ema144 = close_value
        else:
            ema144 = close_value * alpha144 + ema144 * (1.0 - alpha144)

        if ema169 is None:
            ema169 = close_value
        else:
            ema169 = close_value * alpha169 + ema169 * (1.0 - alpha169)

        candles.append(
            {
                "ts": to_ts(dt_value),
                "dt": dt_value,
                "close": close_value,
                "ema144": float(ema144),
                "ema169": float(ema169),
            }
        )
    return candles


def load_asset_price_series(path: Path) -> List[Dict[str, Any]]:
    points: List[Dict[str, Any]] = []
    for row in read_csv_rows(path):
        dt_value = parse_dt(row.get("dt"))
        close_value = parse_float(row.get("close"), default=float("nan"))
        if dt_value is None or math.isnan(close_value):
            continue
        points.append(
            {
                "ts": to_ts(dt_value),
                "close": round_value(close_value, 2),
            }
        )
    return points


def classify_regimes(candles: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not candles:
        return []

    vegas_states: List[str] = []
    regime_states: List[str] = []
    prev_ema144: Optional[float] = None
    prev_ema169: Optional[float] = None
    for index, candle in enumerate(candles):
        ema144 = candle["ema144"]
        ema169 = candle["ema169"]
        upper = max(ema144, ema169)
        lower = min(ema144, ema169)
        close = candle["close"]

        slope144_pct = 0.0 if prev_ema144 in (None, 0) else ((ema144 - prev_ema144) / prev_ema144) * 100.0
        slope169_pct = 0.0 if prev_ema169 in (None, 0) else ((ema169 - prev_ema169) / prev_ema169) * 100.0

        if close > upper:
            vegas_state = "ABOVE"
        elif close < lower:
            vegas_state = "BELOW"
        else:
            vegas_state = "INSIDE"

        vegas_states.append(vegas_state)
        lookback = vegas_states[max(0, index - 5): index + 1]
        above_ct = sum(1 for item in lookback if item == "ABOVE")
        below_ct = sum(1 for item in lookback if item == "BELOW")
        inside_ct = sum(1 for item in lookback if item == "INSIDE")
        cross_ct = sum(1 for prev, curr in zip(lookback, lookback[1:]) if prev != curr)
        flat_slopes = abs(slope144_pct) <= 0.15 and abs(slope169_pct) <= 0.15

        if above_ct >= 4 and slope144_pct > 0.15 and slope169_pct > 0.15:
            regime_state = "TREND_UP"
        elif below_ct >= 4 and slope144_pct < -0.15 and slope169_pct < -0.15:
            regime_state = "TREND_DOWN"
        elif (cross_ct >= 2 and flat_slopes) or (inside_ct >= 3 and flat_slopes):
            regime_state = "RANGE_CHOP"
        else:
            regime_state = "TRANSITION"

        regime_states.append(regime_state)
        prev_ema144 = ema144
        prev_ema169 = ema169

    segments: List[Dict[str, Any]] = []
    start_index = 0
    current = regime_states[0]
    for index in range(1, len(regime_states)):
        if regime_states[index] == current:
            continue
        segments.append(
            {
                "state": current,
                "start_ts": candles[start_index]["ts"],
                "end_ts": candles[index - 1]["ts"],
            }
        )
        current = regime_states[index]
        start_index = index
    segments.append(
        {
            "state": current,
            "start_ts": candles[start_index]["ts"],
            "end_ts": candles[-1]["ts"],
        }
    )
    return segments


def build_equity_curve(trades: List[Dict[str, Any]], start_ts: str) -> List[Dict[str, Any]]:
    grouped: Dict[str, float] = defaultdict(float)
    for trade in trades:
        grouped[trade["exit_ts"]] += trade["net_pnl"]

    equity = INITIAL_EQUITY
    peak = equity
    points: List[Dict[str, Any]] = [
        {
            "ts": start_ts,
            "equity": round_value(equity, 2),
            "drawdown_pct": 0.0,
        }
    ]
    for ts in sorted(grouped):
        equity += grouped[ts]
        peak = max(peak, equity)
        drawdown_pct = ((equity / peak) - 1.0) * 100.0 if peak > 0 else 0.0
        points.append(
            {
                "ts": ts,
                "equity": round_value(equity, 2),
                "drawdown_pct": round_value(drawdown_pct, 4),
            }
        )
    return points


def build_asset_curves(trades: List[Dict[str, Any]], start_ts: str) -> Dict[str, List[Dict[str, Any]]]:
    grouped: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for trade in trades:
        grouped[trade["asset"]][trade["exit_ts"]] += trade["net_pnl"]

    curves: Dict[str, List[Dict[str, Any]]] = {}
    for asset in ("BTC", "ETH", "SOL"):
        running = 0.0
        points = [{"ts": start_ts, "pnl": 0.0}]
        for ts in sorted(grouped[asset]):
            running += grouped[asset][ts]
            points.append({"ts": ts, "pnl": round_value(running, 2)})
        curves[asset] = points
    return curves


def build_asset_stats(trades: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    stats: List[Dict[str, Any]] = []
    for asset in ("BTC", "ETH", "SOL"):
        items = [trade for trade in trades if trade["asset"] == asset]
        if not items:
            continue
        wins = sum(1 for trade in items if trade["net_pnl"] > 0)
        gross_profit = sum(trade["net_pnl"] for trade in items if trade["net_pnl"] > 0)
        gross_loss = -sum(trade["net_pnl"] for trade in items if trade["net_pnl"] < 0)
        profit_factor = gross_profit / gross_loss if gross_loss > 0 else None
        total_pnl = sum(trade["net_pnl"] for trade in items)
        stats.append(
            {
                "asset": asset,
                "trades": len(items),
                "win_rate_pct": round_value(wins / len(items) * 100.0, 2),
                "total_pnl": round_value(total_pnl, 2),
                "avg_pnl": round_value(total_pnl / len(items), 2),
                "profit_factor": round_value(profit_factor, 2) if profit_factor is not None else None,
            }
        )
    return stats


def build_trade_overview(trades: List[Dict[str, Any]]) -> Dict[str, Any]:
    wins = [trade for trade in trades if float(trade["net_pnl"]) > 0]
    losses = [trade for trade in trades if float(trade["net_pnl"]) < 0]
    return {
        "winning_trades": len(wins),
        "winning_pnl": round_value(sum(float(trade["net_pnl"]) for trade in wins), 2),
        "losing_trades": len(losses),
        "losing_pnl": round_value(sum(float(trade["net_pnl"]) for trade in losses), 2),
        "win_rate_pct": round_value((len(wins) / len(trades)) * 100.0, 2) if trades else 0.0,
    }


def load_regime_snapshot() -> Dict[str, Any]:
    if not REGIME_AUDIT_PATH.exists():
        return {}

    wb = load_workbook(REGIME_AUDIT_PATH, read_only=True, data_only=True)

    def load_rows(sheet_name: str) -> List[Dict[str, Any]]:
        ws = wb[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        header = [str(item) if item is not None else "" for item in next(iterator)]
        rows: List[Dict[str, Any]] = []
        for row in iterator:
            rows.append({header[idx]: row[idx] for idx in range(len(header))})
        return rows

    summary_group = load_rows("Summary_Group")
    summary_regime = load_rows("Summary_Regime")
    loss_streaks = load_rows("Loss_Streaks")

    group_rows = [row for row in summary_group if row.get("scope") == "ALL" and row.get("symbol") == "ALL"]
    trend_row = next((row for row in group_rows if row.get("regime_group") == "TREND"), None)
    range_row = next((row for row in group_rows if row.get("regime_group") == "RANGE_OR_TRANSITION"), None)

    top_streak = next((row for row in loss_streaks if row.get("scope") == "ALL"), None)
    regime_rows = [
        row for row in summary_regime
        if row.get("scope") == "ALL" and row.get("symbol") == "ALL"
    ]
    regime_detail = [
        {
            "regime_proxy": str(row.get("regime_proxy") or ""),
            "trades": int(row.get("trades", 0) or 0),
            "win_rate_pct": round_value(float(row.get("win_rate_pct", 0) or 0), 2),
            "total_pnl_usd": round_value(float(row.get("total_pnl_usd", 0) or 0), 2),
            "profit_factor": round_value(float(row.get("profit_factor", 0) or 0), 2),
            "max_drawdown_usd": round_value(float(row.get("max_drawdown_usd", 0) or 0), 2),
        }
        for row in regime_rows
    ]

    return {
        "trend": {
            "trades": int(trend_row.get("trades", 0)) if trend_row else 0,
            "win_rate_pct": round_value(float(trend_row.get("win_rate_pct", 0) or 0), 2) if trend_row else 0.0,
            "total_pnl_usd": round_value(float(trend_row.get("total_pnl_usd", 0) or 0), 2) if trend_row else 0.0,
            "profit_factor": round_value(float(trend_row.get("profit_factor", 0) or 0), 2) if trend_row else 0.0,
        },
        "range_or_transition": {
            "trades": int(range_row.get("trades", 0)) if range_row else 0,
            "win_rate_pct": round_value(float(range_row.get("win_rate_pct", 0) or 0), 2) if range_row else 0.0,
            "total_pnl_usd": round_value(float(range_row.get("total_pnl_usd", 0) or 0), 2) if range_row else 0.0,
            "profit_factor": round_value(float(range_row.get("profit_factor", 0) or 0), 2) if range_row else 0.0,
        },
        "worst_streak": {
            "symbol": str(top_streak.get("symbol") or "") if top_streak else "",
            "trades": int(top_streak.get("trades_in_streak", 0)) if top_streak else 0,
            "total_pnl_usd": round_value(float(top_streak.get("total_pnl_usd", 0) or 0), 2) if top_streak else 0.0,
            "dominant_regime_group": str(top_streak.get("dominant_regime_group") or "") if top_streak else "",
            "dominant_regime_proxy": str(top_streak.get("dominant_regime_proxy") or "") if top_streak else "",
            "start_dt": str(top_streak.get("start_dt") or "") if top_streak else "",
            "end_dt": str(top_streak.get("end_dt") or "") if top_streak else "",
        },
        "regime_detail": regime_detail,
    }


def load_regime_trade_detail() -> List[Dict[str, Any]]:
    if not REGIME_AUDIT_PATH.exists():
        return []

    wb = load_workbook(REGIME_AUDIT_PATH, read_only=True, data_only=True)
    ws = wb["All_Trades"]
    iterator = ws.iter_rows(values_only=True)
    header = [str(item) if item is not None else "" for item in next(iterator)]
    rows: List[Dict[str, Any]] = []
    for raw in iterator:
        row = {header[idx]: raw[idx] for idx in range(len(header))}
        exit_dt = parse_dt(row.get("exit_dt"))
        entry_dt = parse_dt(row.get("entry_dt"))
        if exit_dt is None or entry_dt is None:
            continue
        rows.append(
            {
                "symbol": str(row.get("symbol") or ""),
                "type": str(row.get("type") or ""),
                "side": str(row.get("side") or ""),
                "entry_dt": to_ts(entry_dt),
                "exit_dt": to_ts(exit_dt),
                "hold_bars": int(row.get("hold_bars") or 0),
                "pnl_pct": round_value(float(row.get("pnl_pct") or 0), 4),
                "pnl_usd": round_value(float(row.get("pnl_usd") or 0), 2),
                "regime_proxy": str(row.get("regime_proxy") or ""),
                "regime_group": str(row.get("regime_group") or ""),
            }
        )
    return rows


def compute_monthly_returns(equity_curve: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[int]]:
    month_equity: Dict[Tuple[int, int], float] = {}
    for point in equity_curve:
        dt_value = parse_dt(point["ts"])
        if dt_value is None:
            continue
        month_equity[(dt_value.year, dt_value.month)] = float(point["equity"])

    ordered_keys = sorted(month_equity)
    monthly: List[Dict[str, Any]] = []
    previous_equity: Optional[float] = None
    years = sorted({year for year, _ in ordered_keys})
    for year, month in ordered_keys:
        equity = month_equity[(year, month)]
        if previous_equity is None or previous_equity == 0:
            ret = None
            pnl_usd = None
        else:
            ret = ((equity / previous_equity) - 1.0) * 100.0
            pnl_usd = equity - previous_equity
        monthly.append(
            {
                "year": year,
                "month": month,
                "return_pct": round_value(ret, 2) if ret is not None else None,
                "pnl_usd": round_value(pnl_usd, 2) if pnl_usd is not None else None,
                "end_equity": round_value(equity, 2),
            }
        )
        previous_equity = equity
    return monthly, years


def build_monthly_heatmap(monthly_returns: List[Dict[str, Any]], years: List[int]) -> Dict[str, Any]:
    lookup = {(item["year"], item["month"]): item["return_pct"] for item in monthly_returns}
    matrix: List[List[Optional[float]]] = []
    values: List[float] = []
    for year in years:
        row: List[Optional[float]] = []
        for month in range(1, 13):
            value = lookup.get((year, month))
            row.append(value)
            if value is not None:
                values.append(value)
        matrix.append(row)
    return {
        "years": years,
        "months": ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"],
        "matrix": matrix,
        "min_value": round_value(min(values), 2) if values else -1.0,
        "max_value": round_value(max(values), 2) if values else 1.0,
    }


def build_monthly_summary(monthly_returns: List[Dict[str, Any]]) -> Dict[str, Any]:
    valid = [item for item in monthly_returns if item["pnl_usd"] is not None]
    positive = [item for item in valid if float(item["pnl_usd"]) > 0]
    negative = [item for item in valid if float(item["pnl_usd"]) < 0]
    flat = [item for item in valid if float(item["pnl_usd"]) == 0]

    def month_label(item: Optional[Dict[str, Any]]) -> Optional[str]:
        if item is None:
            return None
        return f"{item['year']}-{int(item['month']):02d}"

    best_month = max(valid, key=lambda item: float(item["pnl_usd"]), default=None)
    worst_month = min(valid, key=lambda item: float(item["pnl_usd"]), default=None)

    return {
        "profitable_months": len(positive),
        "profitable_months_pnl": round_value(sum(float(item["pnl_usd"]) for item in positive), 2),
        "losing_months": len(negative),
        "losing_months_pnl": round_value(sum(float(item["pnl_usd"]) for item in negative), 2),
        "flat_months": len(flat),
        "best_month": {
            "label": month_label(best_month),
            "pnl_usd": round_value(float(best_month["pnl_usd"]), 2) if best_month is not None else None,
            "return_pct": round_value(float(best_month["return_pct"]), 2) if best_month and best_month["return_pct"] is not None else None,
        },
        "worst_month": {
            "label": month_label(worst_month),
            "pnl_usd": round_value(float(worst_month["pnl_usd"]), 2) if worst_month is not None else None,
            "return_pct": round_value(float(worst_month["return_pct"]), 2) if worst_month and worst_month["return_pct"] is not None else None,
        },
    }


def build_monthly_breakdown(trades: List[Dict[str, Any]], monthly_returns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    returns_lookup = {(item["year"], item["month"]): item for item in monthly_returns}
    grouped: Dict[Tuple[int, int], List[Dict[str, Any]]] = defaultdict(list)
    for trade in trades:
        grouped[(trade["exit_dt"].year, trade["exit_dt"].month)].append(trade)

    months: List[Dict[str, Any]] = []
    for year_month in sorted(grouped.keys()):
        year, month = year_month
        items = sorted(grouped[year_month], key=lambda trade: (trade["exit_dt"], trade["asset"]))
        wins = sum(1 for trade in items if float(trade["net_pnl"]) > 0)
        losses = sum(1 for trade in items if float(trade["net_pnl"]) < 0)
        total_pnl = sum(float(trade["net_pnl"]) for trade in items)
        month_return = returns_lookup.get(year_month)

        by_asset: List[Dict[str, Any]] = []
        for asset in ("BTC", "ETH", "SOL"):
            asset_items = [trade for trade in items if trade["asset"] == asset]
            if not asset_items:
                continue
            by_asset.append(
                {
                    "asset": asset,
                    "trade_count": len(asset_items),
                    "pnl_usd": round_value(sum(float(trade["net_pnl"]) for trade in asset_items), 2),
                }
            )

        months.append(
            {
                "label": f"{year}-{month:02d}",
                "year": year,
                "month": month,
                "trade_count": len(items),
                "wins": wins,
                "losses": losses,
                "net_pnl": round_value(total_pnl, 2),
                "return_pct": month_return["return_pct"] if month_return is not None else None,
                "by_asset": by_asset,
                "trades": [
                    {
                        "asset": trade["asset"],
                        "direction": trade["direction"],
                        "type": trade["type"],
                        "entry_ts": trade["entry_ts"],
                        "exit_ts": trade["exit_ts"],
                        "entry_price": round_value(float(trade["entry_price"]), 4),
                        "exit_price": round_value(float(trade["exit_price"]), 4),
                        "gross_pnl": round_value(float(trade["gross_pnl"]), 2),
                        "fees": round_value(float(trade["fees"]), 2),
                        "net_pnl": round_value(float(trade["net_pnl"]), 2),
                    }
                    for trade in items
                ],
            }
        )
    return months


def compute_summary(equity_curve: List[Dict[str, Any]], monthly_returns: List[Dict[str, Any]], trades: List[Dict[str, Any]]) -> Dict[str, Any]:
    first_dt = parse_dt(equity_curve[0]["ts"])
    last_dt = parse_dt(equity_curve[-1]["ts"])
    initial_equity = float(equity_curve[0]["equity"])
    final_equity = float(equity_curve[-1]["equity"])
    total_return_pct = ((final_equity / initial_equity) - 1.0) * 100.0 if initial_equity > 0 else 0.0
    years = ((last_dt - first_dt).total_seconds() / (365.25 * 24.0 * 3600.0)) if first_dt and last_dt else 0.0
    annualized_return_pct = (((final_equity / initial_equity) ** (1.0 / years)) - 1.0) * 100.0 if years > 0 and initial_equity > 0 else 0.0
    max_drawdown_pct = min(float(point["drawdown_pct"]) for point in equity_curve)

    return_values = [item["return_pct"] / 100.0 for item in monthly_returns if item["return_pct"] is not None]
    sharpe = 0.0
    if len(return_values) >= 2:
        mean_return = sum(return_values) / len(return_values)
        variance = sum((value - mean_return) ** 2 for value in return_values) / (len(return_values) - 1)
        std = math.sqrt(variance)
        if std > 0:
            sharpe = (mean_return / std) * math.sqrt(12.0)

    return {
        "initial_equity": round_value(initial_equity, 2),
        "final_equity": round_value(final_equity, 2),
        "total_return_pct": round_value(total_return_pct, 2),
        "annualized_return_pct": round_value(annualized_return_pct, 2),
        "max_drawdown_pct": round_value(max_drawdown_pct, 2),
        "sharpe": round_value(sharpe, 2),
        "trade_count": len(trades),
        "start_ts": equity_curve[0]["ts"],
        "end_ts": equity_curve[-1]["ts"],
    }


def main() -> None:
    trades = load_portfolio_trades()
    candles = load_btc_candles()
    if not trades or not candles:
        raise SystemExit("Missing required CSV inputs for joint backtest overview.")

    start_ts = candles[0]["ts"]
    regimes = classify_regimes(candles)
    equity_curve = build_equity_curve(trades, start_ts)
    asset_curves = build_asset_curves(trades, start_ts)
    asset_stats = build_asset_stats(trades)
    monthly_returns, years = compute_monthly_returns(equity_curve)
    monthly_summary = build_monthly_summary(monthly_returns)
    monthly_breakdown = build_monthly_breakdown(trades, monthly_returns)
    trade_overview = build_trade_overview(trades)
    regime_snapshot = load_regime_snapshot()
    regime_trade_detail = load_regime_trade_detail()
    summary = compute_summary(equity_curve, monthly_returns, trades)
    heatmap = build_monthly_heatmap(monthly_returns, years)

    payload = {
        "meta": {
            "title": "Joint Backtest Overview",
            "subtitle": "Portfolio equity, BTC regime overlay, asset contribution, and monthly return heatmap",
            "source_files": {
                "portfolio": str(PORTFOLIO_PATH.relative_to(ROOT)),
                "btc_candles": str(BTC_CANDLES_PATH.relative_to(ROOT)),
                "eth_candles": str(ETH_CANDLES_PATH.relative_to(ROOT)),
                "sol_candles": str(SOL_CANDLES_PATH.relative_to(ROOT)),
            },
            "position_config": {
                "margin_usd": round_value(TARGET_MARGIN, 2),
                "leverage": round_value(TARGET_LEVERAGE, 2),
                "nominal_usd": round_value(TARGET_NOMINAL, 2),
            },
        },
        "summary": summary,
        "trade_overview": trade_overview,
        "regime_snapshot": regime_snapshot,
        "regime_trade_detail": regime_trade_detail,
        "monthly_summary": monthly_summary,
        "monthly_breakdown": monthly_breakdown,
        "equity_curve": equity_curve,
        "btc_price": [
            {
                "ts": candle["ts"],
                "close": round_value(candle["close"], 2),
            }
            for candle in candles
        ],
        "asset_prices": {
            "BTC": load_asset_price_series(BTC_CANDLES_PATH),
            "ETH": load_asset_price_series(ETH_CANDLES_PATH),
            "SOL": load_asset_price_series(SOL_CANDLES_PATH),
        },
        "regime_segments": regimes,
        "asset_curves": asset_curves,
        "asset_stats": asset_stats,
        "monthly_returns": monthly_returns,
        "monthly_heatmap": heatmap,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[overview] wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
